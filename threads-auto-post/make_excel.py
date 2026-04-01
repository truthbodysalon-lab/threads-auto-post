import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date

def get_latest(logfile):
    entries = []
    with open(logfile) as f:
        for line in f:
            line = line.strip()
            if line:
                entries.append(json.loads(line))
    return entries[-1] if entries else None

def classify(text):
    lines = [l for l in text.split('\n') if l.strip()]
    return '短文' if len(lines) <= 3 else '展開型'

def build_sheet(ws, account, theme, posts):
    headers = ['No', '投稿文', '文字数', '種別']
    header_font = Font(name='Arial', bold=True)
    header_fill = PatternFill('solid', start_color='D9D9D9')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 10
    for i, post in enumerate(posts, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i).font = Font(name='Arial')
        cell_b = ws.cell(row=row, column=2, value=post)
        cell_b.font = Font(name='Arial')
        cell_b.alignment = Alignment(wrap_text=True)
        ws.cell(row=row, column=3, value=f'=LEN(B{row})').font = Font(name='Arial')
        ws.cell(row=row, column=4, value=classify(post)).font = Font(name='Arial')
        line_count = max(1, len([l for l in post.split('\n') if l.strip()]))
        ws.row_dimensions[row].height = max(15, line_count * 15)

truth = get_latest('/Users/mt112/Desktop/threads-auto-post/log_truth.jsonl')
masa = get_latest('/Users/mt112/Desktop/threads-auto-post/log_masa.jsonl')

wb = Workbook()
ws1 = wb.active
ws1.title = '@truth_body_salon'
build_sheet(ws1, truth['account'], truth['theme'], truth['posts'])

ws2 = wb.create_sheet('@masahide_takahashi_')
build_sheet(ws2, masa['account'], masa['theme'], masa['posts'])

today = date.today().strftime('%Y-%m-%d')
out = f'/Users/mt112/Desktop/Threads投稿_{today}.xlsx'
wb.save(out)
print(f'保存完了: {out}')
